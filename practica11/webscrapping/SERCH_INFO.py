import requests
import os
from os import link
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook


# Se crea el libro de excel

def InfoExcel():
    libro = Workbook()
    hoja = libro.active
    hoja1 = libro.create_sheet("NOTICIAS")
    hoja2 = libro.create_sheet("JUGADORES")
    hoja3 = libro.create_sheet("REDES SOCIALES")
    libro.remove_sheet(libro.get_sheet_by_name('Sheet'))
    print("Excel creado")
    num_line = 1
    with open("URL/Noticias.txt") as archivo_txt:
        for line in archivo_txt:
            url_from_txt = re.findall(r'https?://www\.?\w+\.\w+', line)
            print(">URL obtenido de \'Noticias.txt\' >>> ", url_from_txt)

    length = len(url_from_txt)
    for i in range(length):
        url = url_from_txt[i]

    page = requests.get(url)
    soup = BeautifulSoup(page.text, "html.parser")

    # Noticias del club
    print("***** NOTICIAS DEL CLUB *****")
    Noticias = soup.find_all(class_="titulo")
    for noticia in Noticias:
        print(noticia.text.strip())
        hoja1.cell(num_line, 1, noticia.text.strip())
        num_line+=1

    # Jugadores de Rayados
    num_line=1
    print("*** Jugadores de Rayados ***")
    ListaJugadores = soup.find_all('div', {'class': 'texto'},limit=23)
    for jugador in ListaJugadores:
        jug = jugador.find("h4").text.strip()
        posicion = jugador.find("h5").text.strip()
        print(jug)
        print(posicion)
        print()
        hoja2.cell(num_line, 1,jug)
        hoja2.cell(num_line,3,posicion)
        num_line+=1

    # Redes Sociales
    num_line=1
    RedesSociales = soup.find_all("li", {"class": "redes-xs"})
    for redes in RedesSociales:
        red = redes.a["href"]
        print(red)
        hoja3.cell(num_line, 1, red)
        num_line+=1
        
    libro.save("DatosRayados.xlsx")
    